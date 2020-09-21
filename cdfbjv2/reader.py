# coding: utf-8

import random
import logging
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


def load_goods_list(filename):
    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows
    goods_urls = [row[0].value for row in rows if row and row[0] and row[0].value is not None]
    goods_list = [url.split('/')[-1].split('?')[0] for url in goods_urls if url.strip()]

    logger.info('goods list: ----------------------------------------------------------------')
    logger.info(goods_list)
    logger.info('---------------------------------------------------------------------------------')

    return list(set(goods_list))


def load_user_info_dict(filename):
    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows

    user_info_dict = {}

    for row in rows:
        user = row[0].value
        code = row[1].value
        if not user or not code:
            continue

        user = user.strip()
        code = code.strip()

        if '@' in user and user not in user_info_dict.keys():
            user_info_dict[user] = {'email': user, 'email_code': code}

    logger.info('user info dict: ----------------------------------------------------------------')
    logger.info(user_info_dict)
    logger.info('---------------------------------------------------------------------------------')

    return user_info_dict




def load_goods_user_info_v2(filename):
    goods_user_info = load_goods_user_info(filename)
    user_info_dict = {}
    for goods_id, users in goods_user_info.items():
        for user in users:
            if user not in user_info_dict.keys():
                user_info_dict[user] = {}
                user_info_dict[user]['email'] = user
                user_info_dict[user]['password'] = None
                user_info_dict[user]['goods'] = {}
            user_info_dict[user]['goods'][goods_id] = 1
    logger.info('goods user info: ----------------------------------------------------------------')
    logger.info(goods_user_info)
    logger.info('---------------------------------------------------------------------------------')
    logger.info('user info dict: ------------------------------------------------------------------')
    logger.info(user_info_dict)
    logger.info('---------------------------------------------------------------------------------')
    return goods_user_info, user_info_dict


def load_auto_order_goods_user_info(filename):
    user_info_dict = {}
    goods_user_info = {}

    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows

    i = 1
    for row in rows:
        i = i+1
        email = booksheet.cell(row=i, column=3).value
        if email is None or email.strip().isspace() or '@' not in email:
            continue
        email = email.strip()
        user = email

        email_code = None
        # email_code = booksheet.cell(row=i, column=16).value
        # if email_code is not None and not email_code.strip().isspace():
        #     email_code = email_code.strip()
        # else:
        #     email_code = None

        if user in user_info_dict.keys():     # 同一个人过去的数据
            continue

        user_info_dict[user] = {'email': email, 'email_code': email_code, 'goods': {}}

        arr = [4, 7, 10, 13, 16, 19]
        url_order_num_list = [(booksheet.cell(row=i, column=j).value, booksheet.cell(row=i, column=j+1).value,
                               booksheet.cell(row=i, column=j+2).value) for j in arr]
        url_order_num_list = [(x, y, z) for (x, y, z) in url_order_num_list if x is not None]
        for url, order_num, order_times in url_order_num_list:
            try:
                url = url.strip('')
                if 'm.yuegowu.com/goods-detail' not in url:
                    continue
                goods_id = url.split('/')[-1].split('?')[0]

                if order_num is None or not order_num.isnumeric() or int(order_num) <= 1:
                    order_num = 1
                else:
                    order_num = int(order_num)

                if order_times is None or not order_times.isnumeric() or int(order_times) <= 1:
                    order_times = 1
                else:
                    order_times = int(order_times)

                user_info_dict[user]['goods'][goods_id] = (order_num, order_times)

            except Exception as e:
                logger.exception('parse[{0}] user info [{1}:{2}] exception: [{3}].'.format(user, url, order_num, e))

    for user, user_info in user_info_dict.items():
        for goods_id in user_info['goods'].keys():
            if goods_id not in goods_user_info.keys():
                goods_user_info[goods_id] = []
            goods_user_info[goods_id].append(user)

    logger.info('goods user info: ----------------------------------------------------------------')
    logger.info(goods_user_info)
    logger.info('---------------------------------------------------------------------------------')
    return goods_user_info, user_info_dict


def update_auto_order_user_info(filename, user_info_dict):
    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows

    i = 1
    for row in rows:
        i = i+1

        email = booksheet.cell(row=i, column=4).value
        if email is None or email.strip().isspace() or '@' not in email:
            continue
        email = email.strip()
        user = email

        if user not in user_info_dict:
            user_info_dict[user] = {'email': email, 'email_code': None, 'goods': {}}

        if user_info_dict[user]['email_code'] is None:
            email_code = booksheet.cell(row=i, column=5).value
            if email_code is not None and not email_code.strip().isspace():
                email_code = email_code.strip()
            else:
                email_code = None
            user_info_dict[user]['email_code'] = email_code

        wechat = booksheet.cell(row=i, column=3).value
        user_info_dict[user]['wechat'] = wechat

        login_user = booksheet.cell(row=i, column=6).value
        user_info_dict[user]['login_user'] = login_user.strip()

        login_password = booksheet.cell(row=i, column=7).value
        user_info_dict[user]['login_password'] = login_password.strip()

        province = booksheet.cell(row=i, column=8).value
        user_info_dict[user]['province'] = province.strip()

        city = booksheet.cell(row=i, column=9).value
        user_info_dict[user]['city'] = city.strip()

        distinct = booksheet.cell(row=i, column=10).value
        user_info_dict[user]['distinct'] = distinct.strip()

        address = booksheet.cell(row=i, column=11).value
        user_info_dict[user]['address'] = address.strip()

        # flight = booksheet.cell(row=i, column=9).value
        flight = 'SU20{0}'.format(random.choice(range(10)))
        user_info_dict[user]['flight'] = flight.strip()

        arrive_time = '2017-0{0}-0{1} 12:00:00'.format(random.choice(range(1, 10)), random.choice(range(1, 10)))
        # arrive_time = booksheet.cell(row=i, column=10).value
        user_info_dict[user]['arrive_time'] = arrive_time.strip()

        seat = 'L3{0}'.format(random.choice(range(10)))
        # seat = booksheet.cell(row=i, column=11).value
        user_info_dict[user]['seat'] = seat.strip()

        passport = 'G50442{0}96'.format(random.choice(range(10)))
        # passport = booksheet.cell(row=i, column=12).value
        user_info_dict[user]['passport'] = passport.strip()

    logger.info('user info dict: ------------------------------------------------------------------')
    logger.info(user_info_dict)
    logger.info('---------------------------------------------------------------------------------')


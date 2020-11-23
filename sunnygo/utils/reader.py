# coding: utf-8


import logging
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


def load_sys_goods_file(filename):
    """读取管理员权限商品信息表"""

    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows
    goods_urls = [row[0].value for row in rows if row and row[0] and row[0].value is not None]
    goods_list = [url.split('/')[-1].split('?')[0] for url in goods_urls if url.strip()]

    logger.info('load_sys_goods_file: ------------------------------------------------------------')
    logger.info(goods_list)
    logger.info('---------------------------------------------------------------------------------')

    return list(set(goods_list))


def load_sys_user_file(filename):
    """读取管理员权限用户信息表"""

    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows

    sys_user_dict = {}

    for row in rows:
        item = {'user_id': row[3].value, 'email': row[0].value, 'email_code': row[1].value, 'name': row[2].value}

        if item['user_id'] is None:
            logger.info('load sys user file[{0}] wrong user_id format[{1}].'.format(filename, item))
            continue
        elif item['email'] is not None and '@' not in item['email']:
            logger.info('load sys user file[{0}] wrong email format[{1}].'.format(filename, item))
            continue

        item['user_id'] = str(item['user_id']).strip()
        if item['user_id'] in sys_user_dict.keys():
            logger.info('load sys user file[{0}] repeated user_id[{1}].'.format(filename, item))
            continue

        if item['email'] is not None:
            item['email'] = str(item['email']).strip()

        if item['email_code'] is not None:
            item['email_code'] = str(item['email_code']).strip()

        if item['name'] is not None:
            item['name'] = str(item['name']).strip()

        sys_user_dict[item['user_id']] = item

    logger.info('load_sys_user_file: ----------------------------------------------------------------')
    logger.info(sys_user_dict)
    logger.info('---------------------------------------------------------------------------------')

    return sys_user_dict


def load_zmhttp_city_code_dict(filename):
    """读取zmhttp的城市对应码，以方便获取指定城市的ip"""

    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows

    city_code_dict = {}

    for row in rows:
        city_name = row[1].value
        city_code = row[2].value
        if city_name and city_code:
            city_code_dict[city_name.strip()] = str(city_code).strip()

    return city_code_dict

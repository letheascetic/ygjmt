# coding: utf-8

from openpyxl import load_workbook


def load_goods_user_info(filename):
    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    # columns = booksheet.columns

    content = {}
    i = 1
    for row in rows:
        i = i+1
        email = booksheet.cell(row=i, column=3).value
        if email is None or email.strip().isspace() or '@' not in email:
            continue
        email = email.strip()
        goods_ids = []
        if email in content.keys():     # 同一个人过去的数据
            # goods_ids = content[email]
            continue

        urls = [booksheet.cell(row=i, column=j).value for j in range(4, 10)]
        urls = [url for url in urls if url and url.strip()]
        for url in urls:
            try:
                url = [u for u in url.split(' ') if 'goods-detail' in u]
                for u in url:
                    goods_id = u.split('goods-detail/')[-1].split('/')[0]
                    if goods_id not in goods_ids:
                        goods_ids.append(goods_id)
            except Exception as e:
                print('parse[{0}] url[{1}] exception: [{2}].'.format(email, url, e))
        content[email] = goods_ids

    print('---------------------------------------------------------------------------------')
    i = 0
    for email in content.keys():
        i = i + 1
        print('[{0}] [{1}]: [{2}].'.format(i, email, content[email]))
    print('---------------------------------------------------------------------------------')

    goods_user_dict = {}
    for user, goods_ids in content.items():
        for goods_id in goods_ids:
            if goods_id not in goods_user_dict.keys():
                goods_user_dict[goods_id] = set()
            goods_user_dict[goods_id].add(user)

    print('---------------------------------------------------------------------------------')
    i = 0
    for goods_id in goods_user_dict.keys():
        i = i + 1
        print('[{0}] [{1}]: [{2}].'.format(i, goods_id, goods_user_dict[goods_id]))
    print('---------------------------------------------------------------------------------')

    return goods_user_dict


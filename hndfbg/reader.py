# coding: utf-8

import logging
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


def load_goods_user_info(filename):
    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    # columns = booksheet.columns

    content = {}
    i = 1
    for row in rows:
        i = i+1
        email = booksheet.cell(row=i, column=3).value
        if email is None or email.strip().isspace() or '@' not in email:
            continue
        email = email.strip()
        goods_ids = []
        if email in content.keys():
            # goods_ids = content[email]
            continue

        urls = [booksheet.cell(row=i, column=j).value for j in range(4, 10)]
        urls = [url for url in urls if url and url.strip()]

        for url in urls:
            try:
                url = [u for u in url.split(' ') if 'goods_id' in u]
                for u in url:
                    goods_ids.append(u.split('&')[0].split('=')[-1])
            except Exception as e:
                logger.exception('parse[{0}] url[{1}] exception: [{2}].'.format(email, url, e))
        content[email] = goods_ids

    logger.info('---------------------------------------------------------------------------------')
    i = 0
    for email in content.keys():
        i = i + 1
        logger.info('[{0}] [{1}]: [{2}].'.format(i, email, content[email]))
    logger.info('---------------------------------------------------------------------------------')

    goods_user_dict = {}
    for user, goods_ids in content.items():
        for goods_id in goods_ids:
            if goods_id not in goods_user_dict.keys():
                goods_user_dict[goods_id] = set()
            goods_user_dict[goods_id].add(user)

    logger.info('---------------------------------------------------------------------------------')
    i = 0
    for goods_id in goods_user_dict.keys():
        i = i + 1
        logger.info('[{0}] [{1}]: [{2}].'.format(i, goods_id, goods_user_dict[goods_id]))
    logger.info('---------------------------------------------------------------------------------')

    return goods_user_dict


def load_goods_user_info_v3(filename):
    user_info_dict = {}
    goods_user_info = {}

    workbook = load_workbook(filename)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    rows = booksheet.rows

    i = 1
    for row in rows:
        i = i+1
        email = booksheet.cell(row=i, column=3).value

        if email is None:
            continue
        elif email.strip().isspace() or '@' not in email:
            logger.warning('error email format: [{0}].'.format(email))
            continue
        email = email.strip()
        user = email

        email_code = booksheet.cell(row=i, column=11).value
        if email_code is not None and not email_code.strip().isspace():
            email_code = ''.join(email_code.strip().split(' '))
        else:
            email_code = None

        if user in user_info_dict.keys():     # 同一个人过去的数据
            continue

        user_info_dict[user] = {'wechat': None, 'email': email, 'email_code': email_code, 'password': None, 'goods': {}}
        wechat = booksheet.cell(row=i, column=10).value
        if wechat is not None and not wechat.strip().isspace():
            user_info_dict[user]['wechat'] = wechat.strip()
        else:
            user_info_dict[user]['wechat'] = None

        z = [4, 5, 6, 7, 8, 9]
        url_order_num_list = [(booksheet.cell(row=i, column=j).value, 1) for j in z]
        url_order_num_list = [(x, y) for (x, y) in url_order_num_list if x is not None]
        for url, order_num, in url_order_num_list:
            try:
                url = url.split('&')[0]
                if 'm.hndfbg.com/goods/detail?goods_id=' not in url:
                    logger.warning('error goods url[{0}] for user[{1}].'.format(url, email))
                    continue
                goods_id = url.split('=')[-1]

                user_info_dict[user]['goods'][goods_id] = order_num

            except Exception as e:
                logger.exception('parse[{0}] user info [{1}:{2}] exception: [{3}].'.format(user, url, order_num, e))

    for user, user_info in user_info_dict.items():
        for goods_id in user_info['goods'].keys():
            if goods_id not in goods_user_info.keys():
                goods_user_info[goods_id] = []
            goods_user_info[goods_id].append(user)

    logger.info('goods user info: ----------------------------------------------------------------')
    logger.info(goods_user_info)
    logger.info('---------------------------------------------------------------------------------')
    logger.info('user info dict: ------------------------------------------------------------------')
    logger.info(user_info_dict)
    logger.info('---------------------------------------------------------------------------------')
    return goods_user_info, user_info_dict

